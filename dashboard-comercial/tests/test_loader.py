# dashboard-comercial/tests/test_loader.py
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import pytest
import tempfile
import openpyxl
from loader import carregar_excel, carregar_config

def criar_excel_minimo(path):
    """Cria um Excel mínimo válido para testes."""
    wb = openpyxl.Workbook()
    # Aba EMPRESAS
    ws_emp = wb.active
    ws_emp.title = "EMPRESAS"
    ws_emp.append(["ID", "Nome da Empresa", "CNPJ", "Segmento", "Cidade",
                   "Estado", "Nome Contato", "Cargo", "Telefone", "E-mail",
                   "Temperatura", "Porte", "Origem Lead", "Observações Gerais", "Ativo?"])
    ws_emp.append(["EMP001", "Empresa Teste", None, "Serviços", "SP", "SP",
                   "João", "Diretor", None, None, "Quente", "Médio", None, None, "SIM"])
    # Aba VISITAS
    ws_vis = wb.create_sheet("VISITAS")
    ws_vis.append(["Data", "ID Empresa", "Etapa", "Visita Realizada?",
                   "Procuração Assinada?", "Proposta Enviada?", "Contrato Fechado?",
                   "Duração (min)", "Resultado", "Interesse (1-5)", "Próxima Ação",
                   "Data Próx. Ação", "Obstáculo", "Relatório Resumido", "Registrado por"])
    from datetime import date
    ws_vis.append([date(2026, 3, 1), "EMP001", "1ª Visita", "SIM",
                   "NÃO", "NÃO", "NÃO", 30, "Positivo", 4, None, None, None, None, None])
    # Aba CONFIG
    ws_cfg = wb.create_sheet("CONFIG")
    ws_cfg.append(["Parâmetro", "Valor", "Descrição"])
    for row in [
        ["nome_agente", "Agente Teste", ""],
        ["sla_quente_dias", 30, ""],
        ["sla_normal_dias", 90, ""],
        ["ano_filtro", 2026, ""],
        ["empresa_nome", "Empresa X", ""],
        ["cor_primaria", "#001b47", ""],
        ["cor_acento", "#af946c", ""],
        ["alerta_amarelo_pct", 75, ""],
        ["mostrar_cnpj", "SIM", ""],
        ["mostrar_telefone", "NÃO", ""],
    ]:
        ws_cfg.append(row)
    wb.save(path)


def test_carregar_excel_retorna_tres_dataframes():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    criar_excel_minimo(path)
    empresas, visitas, config = carregar_excel(path)
    assert len(empresas) == 1
    assert len(visitas) == 1
    assert config["nome_agente"] == "Agente Teste"


def test_carregar_excel_arquivo_inexistente():
    with pytest.raises(FileNotFoundError):
        carregar_excel("nao_existe.xlsx")


def test_carregar_config_valores_numericos():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    criar_excel_minimo(path)
    _, _, config = carregar_excel(path)
    assert isinstance(config["sla_quente_dias"], int)
    assert isinstance(config["sla_normal_dias"], int)
    assert isinstance(config["alerta_amarelo_pct"], int)
    assert config["sla_quente_dias"] == 30


def test_carregar_config_fallback_campos_vazios():
    wb = openpyxl.Workbook()
    ws_emp = wb.active
    ws_emp.title = "EMPRESAS"
    ws_emp.append(["ID", "Nome da Empresa", "CNPJ", "Segmento", "Cidade",
                   "Estado", "Nome Contato", "Cargo", "Telefone", "E-mail",
                   "Temperatura", "Porte", "Origem Lead", "Observações Gerais", "Ativo?"])
    ws_vis = wb.create_sheet("VISITAS")
    ws_vis.append(["Data", "ID Empresa", "Etapa", "Visita Realizada?",
                   "Procuração Assinada?", "Proposta Enviada?", "Contrato Fechado?",
                   "Duração (min)", "Resultado", "Interesse (1-5)", "Próxima Ação",
                   "Data Próx. Ação", "Obstáculo", "Relatório Resumido", "Registrado por"])
    ws_cfg = wb.create_sheet("CONFIG")
    ws_cfg.append(["Parâmetro", "Valor", "Descrição"])
    ws_cfg.append(["nome_agente", None, ""])   # vazio
    ws_cfg.append(["empresa_nome", None, ""])  # vazio
    ws_cfg.append(["sla_quente_dias", 30, ""])
    ws_cfg.append(["sla_normal_dias", 90, ""])
    ws_cfg.append(["ano_filtro", 2026, ""])
    ws_cfg.append(["cor_primaria", "#001b47", ""])
    ws_cfg.append(["cor_acento", "#af946c", ""])
    ws_cfg.append(["alerta_amarelo_pct", 75, ""])
    ws_cfg.append(["mostrar_cnpj", "SIM", ""])
    ws_cfg.append(["mostrar_telefone", "NÃO", ""])
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    wb.save(path)
    _, _, config = carregar_excel(path)
    assert config["nome_agente"] == "Agente"
    assert config["empresa_nome"] == "Minha Empresa"
